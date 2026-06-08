import time
import json
import os
import datetime
import openpyxl
from collections import defaultdict
from fastapi import FastAPI, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware

from sqlalchemy import text
from .database import engine, Base, get_db
from . import crud
from .routers import auth, apiaries, colonies, traits, stats, sync, researcher, admin

# Create database tables
Base.metadata.create_all(bind=engine)

def sync_schema_columns(db):
    """실제 상용 DB(PostgreSQL 등)에 모델 변경 사항(컬럼 추가 등)을 동적으로 반영하는 경량화 마이그레이션 가드"""
    is_postgres = "postgresql" in str(db.bind.url)
    
    if is_postgres:
        statements = [
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS farm_name VARCHAR(100);",
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS full_name VARCHAR(100);",
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS phone VARCHAR(50);",
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS experience_years INTEGER;",
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS queen_species VARCHAR(50);",
            "ALTER TABLE apiaries ADD COLUMN IF NOT EXISTS owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE;",
            "ALTER TABLE colonies ADD COLUMN IF NOT EXISTS mother_colony_id INTEGER REFERENCES colonies(id) ON DELETE SET NULL;",
            "ALTER TABLE colonies ADD COLUMN IF NOT EXISTS sample_id VARCHAR(100);",
            "ALTER TABLE colonies ADD COLUMN IF NOT EXISTS is_pore_c BOOLEAN DEFAULT FALSE;",
            "ALTER TABLE colonies ADD COLUMN IF NOT EXISTS queen_species VARCHAR(50);",
            "ALTER TABLE trait_records ADD COLUMN IF NOT EXISTS vsh_rate DOUBLE PRECISION DEFAULT 0.0;",
            "ALTER TABLE trait_records ADD COLUMN IF NOT EXISTS hygienic_rate DOUBLE PRECISION DEFAULT 0.0;"
        ]
    else:
        statements = [
            "ALTER TABLE users ADD COLUMN farm_name VARCHAR(100);",
            "ALTER TABLE users ADD COLUMN full_name VARCHAR(100);",
            "ALTER TABLE users ADD COLUMN phone VARCHAR(50);",
            "ALTER TABLE users ADD COLUMN experience_years INTEGER;",
            "ALTER TABLE users ADD COLUMN queen_species VARCHAR(50);",
            "ALTER TABLE apiaries ADD COLUMN owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE;",
            "ALTER TABLE colonies ADD COLUMN mother_colony_id INTEGER REFERENCES colonies(id) ON DELETE SET NULL;",
            "ALTER TABLE colonies ADD COLUMN sample_id VARCHAR(100);",
            "ALTER TABLE colonies ADD COLUMN is_pore_c BOOLEAN DEFAULT FALSE;",
            "ALTER TABLE colonies ADD COLUMN queen_species VARCHAR(50);",
            "ALTER TABLE trait_records ADD COLUMN vsh_rate FLOAT DEFAULT 0.0;",
            "ALTER TABLE trait_records ADD COLUMN hygienic_rate FLOAT DEFAULT 0.0;"
        ]
        
    for stmt in statements:
        try:
            db.execute(text(stmt))
            db.commit()
        except Exception as e:
            db.rollback()
            err_msg = str(e).lower()
            if "duplicate" in err_msg or "already exists" in err_msg:
                pass
            else:
                print(f"[Schema Migration Warning] Failed to run statement '{stmt}': {e}")

app = FastAPI(
    title="K-BEE BANK Central Beekeeping API",
    description="PhD Beekeeping Breeding Record & Telemetry Sync Hub",
    version="2.0.0"
)

# Initialize app state cache
app.state.sampling_cache = {}

# Enable CORS for Next.js web (3000) and Flutter devices/emulators
app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=r"https?://.*",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# API Rate Limiter 미들웨어 장착 (봇 및 DDoS 무차별 대입 차단)
class RateLimitMiddleware(BaseHTTPMiddleware):
    def __init__(self, app, limit: int = 120, window: int = 60):
        super().__init__(app)
        self.limit = limit
        self.window = window
        self.requests = defaultdict(list)

    async def dispatch(self, request: Request, call_next):
        if request.url.path.startswith("/api/v1/auth") or request.method in ["POST", "PUT", "DELETE"]:
            ip = request.client.host if request.client else "127.0.0.1"
            now = time.time()
            self.requests[ip] = [t for t in self.requests[ip] if now - t < self.window]
            if len(self.requests[ip]) >= self.limit:
                return Response(
                    content=json.dumps({"detail": "너무 많은 요청이 발생했습니다. 1분 후에 다시 시도해주세요."}),
                    status_code=429,
                    media_type="application/json"
                )
            self.requests[ip].append(now)
        return await call_next(request)

app.add_middleware(RateLimitMiddleware, limit=120, window=60)

# Auto-seed database on startup (Neon schema migrations synced)
@app.on_event("startup")
def startup_event():
    db = next(get_db())
    try:
        sync_schema_columns(db)
        print("Schema columns synced successfully.")
    except Exception as e:
        print(f"Failed to sync schema columns on startup: {str(e)}")
    crud.seed_database_if_empty(db)

    # Pre-load sampling TSV master databases to memory cache using thread-safe DataManager Singleton
    try:
        from .routers.researcher import BeekeepingDataManager
        manager = BeekeepingDataManager()
        manager.load_data()
        app.state.wgs_cache = manager
        print("Successfully loaded beekeeping TSV dataset to app.state.wgs_cache.")
    except Exception as e:
        print(f"Failed to load beekeeping TSV dataset to memory cache: {str(e)}")

    # On-memory caching of the beekeeping sampling spreadsheet
    excel_path = os.path.join(os.path.dirname(__file__), "data", "샘플링최종.xlsx")
    try:
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
            cache_data = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    cache_data[sheet_name] = []
                    continue
                headers = [str(h) for h in rows[0]]
                sheet_rows = []
                for r in rows[1:]:
                    if any(x is not None for x in r):  # skip blank rows
                        row_dict = {}
                        for col_idx, val in enumerate(r):
                            if col_idx < len(headers):
                                key = headers[col_idx]
                                if isinstance(val, (datetime.datetime, datetime.date)):
                                    row_dict[key] = val.strftime("%Y-%m-%d")
                                else:
                                    row_dict[key] = val
                        sheet_rows.append(row_dict)
                cache_data[sheet_name] = sheet_rows
            app.state.sampling_cache = cache_data
            print("Successfully loaded beekeeping Excel dataset to memory cache.")
        else:
            print(f"Warning: Beekeeping sampling spreadsheet not found at {excel_path}")
    except Exception as e:
        print(f"Failed to load beekeeping Excel sheet into memory: {str(e)}")

@app.get("/")

def read_root():
    return {"name": "K-BEE BANK Central API Server", "status": "Online", "version": "2.0.0"}

# Mount APIRouters
app.include_router(auth.router)
app.include_router(apiaries.router)
app.include_router(colonies.router)
app.include_router(traits.router)
app.include_router(stats.router)
app.include_router(sync.router)
app.include_router(researcher.router)
app.include_router(admin.router)
